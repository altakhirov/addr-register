from django.db import transaction
from django.shortcuts import get_object_or_404

from celery import shared_task
from openpyxl import load_workbook

from addr_register.settings import BASE_DIR
from index.models import FilesToParse
from index.parser import get_community_id, get_street_id, bind_street_with_community, create_house, get_district_id
from main.models import House, Community, Street, Apartment


@shared_task
def xlsx_parser(pk):
    instance = get_object_or_404(FilesToParse, pk=pk)
    wb_obj = load_workbook(filename=instance.file, read_only=True)
    with_flats = instance.with_flats
    sheet_obj = wb_obj.active
    house_list = House.objects.all()
    updated_community_codes = []
    updated_streets_codes = []
    houses_to_update = []
    houses_to_create = []
    i = 0
    for row in sheet_obj.rows:
        if i == 0:
            i = 1
            continue
        i += 1
        data = {
            'cadaster_code': row[1].value,
            'district_oz': row[2].value,
            'district_ru': row[3].value,
            'community_oz': row[4].value,
            'community_ru': row[5].value,
            'community_code': row[6].value,
            'street_oz': row[7].value,
            'street_ru': row[8].value,
            'street_code': row[9].value,
            'extra_oz': row[10].value,
            'extra_ru': row[11].value,
            'number': row[12].value,
            'longitude': row[13].value,
            'latitude': row[14].value,
            'street_type': row[15].value,
            'with_flats': with_flats

        }
        if data.get('community_code') in updated_community_codes:
            try:
                community_id = Community.objects.get(code=data.get('community_code')).id
            except (Community.MultipleObjectsReturned, Community.DoesNotExist):
                community_id = get_community_id(data)
                updated_community_codes.append(data.get('community_code'))
        else:
            community_id = get_community_id(data)
            updated_community_codes.append(data.get('community_code'))
        if data.get('street_code') in updated_streets_codes:
            try:
                street_id = Street.objects.get(code=data.get('street_code'), district_id=get_district_id(data)).id
            except Street.DoesNotExist:
                street_id = get_street_id(data)
                updated_streets_codes.append(data.get('street_code'))
        else:
            street_id = get_street_id(data)
            updated_streets_codes.append(data.get('street_code'))
        try:
            house = house_list.get(cadaster_code=data.get('cadaster_code'))
            with transaction.atomic():
                house.street_id = street_id
                house.community_id = community_id
                house.number = data.get('number')
                house.latitude = data.get('latitude')
                house.longitude = data.get('longitude')
                house.with_flats = with_flats
                # house.save()
                houses_to_update.append(house)
                bind_street_with_community(street_id, community_id)
        except House.MultipleObjectsReturned:
            houses_with_same_cadaster_code = house_list.filter(cadaster_code=data.get('cadaster_code'))
            with transaction.atomic():
                houses_with_same_cadaster_code.delete()
                houses_to_create.append(create_house(data))
        except House.DoesNotExist:
            with transaction.atomic():
                houses_to_create.append(create_house(data))
                bind_street_with_community(street_id, community_id)
        except Exception as e:
            instance.errors += f"Ошибка: Строка {i} ({data.get('cadaster_code')}) - {str(e)}"
        if i % 1000 == 0:
            if len(houses_to_create) > 0:
                House.objects.bulk_create(houses_to_create)
            House.objects.bulk_update(houses_to_update, ['community_id', 'street_id', 'number', 'cadaster_code',
                                                         'latitude', 'longitude', 'with_flats'])
            houses_to_create = []
            houses_to_update = []
    wb_obj.close()
    if houses_to_create:
        House.objects.bulk_create(houses_to_create)
    if houses_to_update:
        House.objects.bulk_update(houses_to_update, ['community_id', 'street_id', 'number', 'cadaster_code',
                                                 'latitude', 'longitude', 'with_flats'])
    instance.in_progress = False
    instance.parsed = True
    instance.save()


@shared_task
def fucking_parser():
    # fucking ijs
    ijs_filename = f'{BASE_DIR}/ijs_domoteka.xlsx'
    wb_obj = load_workbook(filename=ijs_filename, read_only=True)
    sheet_obj = wb_obj.active
    houses_to_update = []
    house_errors = []
    i = 0
    for row in sheet_obj.rows:
        if i == 0:
            i = 1
            continue
        i += 1
        data = {
            'cadaster_code': row[1].value,
            'total_area': row[15].value,
            'building_area': row[16].value,
            'living_area': row[17].value
        }
        try:
            house = House.objects.get(cadaster_code=data.get('cadaster_code'))
            house.total_area = data.get('total_area')
            house.building_area = data.get('building_area')
            house.living_area = data.get('living_area')
            houses_to_update.append(house)
        except Exception as e:
            val = f"{data.get('cadaster_code') + '-' + str(e)}"
            house_errors.append(val)
        if i % 1000 == 0:
            House.objects.bulk_update(houses_to_update, ['total_area', 'building_area', 'living_area'])
            houses_to_update = []
    if len(houses_to_update):
        House.objects.bulk_update(houses_to_update, ['total_area', 'building_area', 'living_area'])
        houses_to_update = []
    wb_obj.close()
    FilesToParse.objects.create(
            parsed=True,
            file=ijs_filename,
            errors=house_errors,
            with_flats=False
        )

    # fucking kv
    kv_filename = f'{BASE_DIR}/kv_domoteka.xlsx'
    wb_obj = load_workbook(filename=kv_filename, read_only=True)
    sheet_obj = wb_obj.active
    apartments_errors = []
    i = 0
    for row in sheet_obj.rows:
        if i == 0:
            i = 1
            continue
        i += 1
        data = {
            'cadaster_code': row[1].value,
            'year': row[23].value,
            'floors': row[24].value,
            'wall_type': row[25].value
        }
        if not data.get('cadaster_code'):
            return
        try:
            apartment = Apartment.objects.get(cadaster_code=data.get('cadaster_code'))
            house = apartment.house
            house.year = data.get('year')
            if data.get('wall_type') == 'Бетон':
                house.wall_type = 'concrete'
            elif data.get('wall_type') == 'Пишиқ ғишт':
                house.wall_type = 'baked_brick'
            elif data.get('wall_type') == 'Темирбетонли (панелли)':
                house.wall_type = 'reinforced_concrete'
            elif data.get('wall_type') == 'Ёғоч девор':
                house.wall_type = 'wooden'
            if house not in houses_to_update:
                houses_to_update.append(house)
        except Exception as e:
            val = f"{data.get('cadaster_code') + '-' + str(e)}"
            apartments_errors.append(val)
        if i % 1000 == 0:
            House.objects.bulk_update(houses_to_update, ['year', 'wall_type'])
    if len(houses_to_update):
        House.objects.bulk_update(houses_to_update, ['year', 'wall_type'])
    wb_obj.close()
    FilesToParse.objects.create(
            parsed=True,
            file=kv_filename,
            errors=apartments_errors,
            with_flats=True
        )


# @shared_task
# def apartment_xlsx_parser():
    # filename = f'{BASE_DIR}/apartments.xlsx'
    # wb_obj = load_workbook(filename=filename, read_only=True)
    # sheet_obj = wb_obj.active
    # apartments_to_create = []
    # multiple_houses_in_cadaster_code = []
    # house_doesnt_exists = []
    # exceptions = []
    # i = 0
    # for row in sheet_obj.rows:
    #     if i == 0:
    #         i = 1
    #         continue
    #     i += 1
    #     data = {
    #         'cadaster_code': row[3].value,
    #         'number': row[6].value,
    #         'rooms': row[10].value,
    #         'house_cadaster_code': row[11].value,
    #         'entrance': row[12].value,
    #         'floor': row[13].value,
    #         'longitude': row[16].value,
    #         'latitude': row[17].value
    #     }
    #     if not data.get('house_cadaster_code'):
    #         return
    #     try:
    #         house = House.objects.get(cadaster_code=data.get('house_cadaster_code'))
    #         apartment = Apartment(
    #             house_id=house.id,
    #             number=data.get('number'),
    #             cadaster_code=data.get('cadaster_code'),
    #             rooms=data.get('rooms'),
    #             entrance=data.get('entrance'),
    #             floor=data.get('floor'),
    #             latitude=data.get('latitude'),
    #             longitude=data.get('longitude')
    #         )
    #         apartments_to_create.append(apartment)
    #     except House.MultipleObjectsReturned:
    #         if data.get('house_cadaster_code') not in multiple_houses_in_cadaster_code:
    #             multiple_houses_in_cadaster_code.append(data.get('house_cadaster_code'))
    #     except House.DoesNotExist:
    #         if data.get('house_cadaster_code') not in house_doesnt_exists:
    #             house_doesnt_exists.append(data.get('house_cadaster_code'))
    #     except Exception as e:
    #         exceptions.append(f"Строка {i}, {data.get('house_cadaster_code')} - {str(e)}")
    #     if i % 1000 == 0:
    #         Apartment.objects.bulk_create(apartments_to_create)
    #         apartments_to_create = []
    # wb_obj.close()
    # if apartments_to_create:
    #     Apartment.objects.bulk_create(apartments_to_create)
    # errors = ''
    # if exceptions:
    #     errors += 'Ошибки:\n'
    #     errors += f"{exceptions}\n"
    #     errors += '=============\n'
    # if multiple_houses_in_cadaster_code:
    #     errors += 'Следующие кадастровые коды привязаны более чем к одному дому:\n'
    #     errors += f"{multiple_houses_in_cadaster_code}\n"
    #     errors += '=============\n'
    # if house_doesnt_exists:
    #     errors += 'Следующие кадастровые коды не привязаны ни к одному дому:\n'
    #     errors += f"{house_doesnt_exists}\n"
    #     errors += '=============\n'
    # FilesToParse.objects.create(
    #     parsed=True,
    #     file=filename,
    #     errors=errors,
    #     with_flats=True
    # )
    # print('Заполняем таблицу частных домов..')

